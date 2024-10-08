# ler a alterar dados em uma planilha do Excel
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando a planilha
workbook: Workbook = load_workbook(WORKBOOK_PATH)
# worksheet: Worksheet = workbook.active # type:ignore

# Nome pra a planilha
sheet_name = 'Minha planilha'
# Selecionando a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2): # type: ignore
    for cell in row:
        print(cell.value,end='\t')
        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2,23)
    print()

# print(worksheet['B3'].value)
# worksheet['B3'].value = 14

workbook.save(WORKBOOK_PATH) # para salvar as alterações